import csv
import random
from pathlib import Path
import openpyxl
from tkinter import messagebox

BASE_DIR = Path(__file__).resolve().parent

# 個数合計の下限と上限
TOTAL_MIN_QTY = 7
TOTAL_MAX_QTY = 9

# 1つの商品番号に対する個数の下限と上限
PROD_MIN_QTY = 1
PROD_MAX_QTY = 3

# フォーマット依存項目
CARD_QTY = 24 # 注文表の枚数
COL_COUNT = 4 # 注文表の横の数
ROW_START = 4 # 開始座標(行)
COL_START = 2 # 開始座標(列)
ROW_SPACING = 11 # セル間隔(縦)
COL_SPACING = 4 # セル間隔(横)
ITEM_TYPE_COUNT = 6 # ピックする商品番号の数

try:
    with open(BASE_DIR / "product.csv", mode="r", encoding="utf_8") as f:
        reader = csv.reader(f)
        lst = list(reader)
except FileNotFoundError:
    messagebox.showwarning("エラー", "商品データ(product.csv)が見つかりません")
    sys.exit()

for i in range(len(lst)):
    lst[i][1] = int(lst[i][1])

stock = dict(lst)

# 出力座標リスト作成
c_list = []

for i in range(CARD_QTY):
    row_block = i // COL_COUNT
    col_block = i % COL_COUNT

    row = ROW_START + row_block * ROW_SPACING
    col = COL_START + col_block * COL_SPACING

    c_list.append([row, col])

# 出力テンプレート呼び出し
try:
    wb = openpyxl.load_workbook(BASE_DIR / "order_sheet_template.xlsx")
    ws = wb.worksheets[0]
except FileNotFoundError:
    messagebox.showwarning("エラー", "テンプレート(order_sheet_template.xlsx)が見つかりません")
    sys.exit()

# 24回繰り返し
for sheet in range(CARD_QTY):

    # 商品番号更新(0を省く)
    products = [k for k, v in stock.items() if v > 0]
    
    order = random.sample(products, ITEM_TYPE_COUNT)

    total_number = random.randint(TOTAL_MIN_QTY, TOTAL_MAX_QTY)

    quantities = []

    while sum(quantities) != total_number:
        quantities.clear()
        for item in order:
            max_qty = min(PROD_MAX_QTY, stock[item])
            qty = random.randint(PROD_MIN_QTY, max_qty)
            quantities.append(qty)

    # 出力
    r = c_list[sheet][0]
    c = c_list[sheet][1]
    for i in range(ITEM_TYPE_COUNT):
        ws.cell(r + i, c).value = order[i]
        ws.cell(r + i, c + 1).value = quantities[i]

    # 在庫更新
    for i in range(ITEM_TYPE_COUNT):
        stock[order[i]] -= quantities[i]

wb.save(BASE_DIR / "order_sheet.xlsx")



